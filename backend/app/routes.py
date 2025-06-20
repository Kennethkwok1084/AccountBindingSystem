from datetime import datetime

import pandas as pd
from flask import Blueprint, jsonify, request, send_file
from openpyxl import load_workbook
from tasks.release import auto_release
from werkzeug.utils import secure_filename

from . import db
from .models import Account, BindingLog
from .utils import add_log

bp = Blueprint("api", __name__)


@bp.route("/ping")
def ping():
    return {"message": "pong"}


@bp.route("/accounts/import", methods=["POST"])
def import_accounts():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "no file"}), 400

    filename = secure_filename(file.filename)
    wb = load_workbook(file)
    ws = wb.active
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        username, password = row[:2]
        if not username or not password:
            continue
        if not Account.query.filter_by(username=username).first():
            account = Account(username=username, password=password)
            db.session.add(account)
            imported += 1
    db.session.commit()
    return jsonify({"imported": imported})


@bp.route("/accounts/bind", methods=["POST"])
def bind_account():
    data = request.get_json() or {}
    username = data.get("username")
    student_id = data.get("student_id")
    if not username or not student_id:
        return jsonify({"error": "missing data"}), 400
    account = Account.query.filter_by(username=username).first()
    if not account:
        return jsonify({"error": "account not found"}), 404
    if account.is_bound:
        return jsonify({"error": "already bound"}), 400
    account.is_bound = True
    account.student_id = student_id
    account.bind_time = datetime.utcnow()
    db.session.add(account)
    db.session.commit()
    add_log(username, student_id, "bind")
    return jsonify({"status": "ok"})


@bp.route("/accounts", methods=["GET"])
def list_accounts():
    """Return account list with optional filters."""
    query = Account.query

    bound = request.args.get("bound")
    if bound is not None:
        if bound.lower() == "true":
            query = query.filter_by(is_bound=True)
        elif bound.lower() == "false":
            query = query.filter_by(is_bound=False)

    q = request.args.get("q")
    if q:
        pattern = f"%{q}%"
        from sqlalchemy import or_

        query = query.filter(
            or_(Account.username.ilike(pattern), Account.student_id.ilike(pattern))
        )

    total = query.count()
    page = int(request.args.get("page", 1))
    size = int(request.args.get("size", 10))
    accounts = query.order_by(Account.id).offset((page - 1) * size).limit(size).all()
    data = [
        {
            "username": a.username,
            "is_bound": a.is_bound,
            "student_id": a.student_id,
            "bind_time": a.bind_time.isoformat() if a.bind_time else None,
        }
        for a in accounts
    ]
    return jsonify({"total": total, "items": data})


@bp.route("/auto-release", methods=["POST"])
def trigger_auto_release():
    days = 32
    if request.is_json:
        days = int(request.get_json().get("days", 32))
    released = auto_release(days)
    return jsonify({"released": released})


@bp.route("/logs", methods=["GET"])
def list_logs():
    """Return binding logs with optional date filters."""
    query = BindingLog.query

    start = request.args.get("start")
    if start:
        query = query.filter(BindingLog.bind_time >= datetime.fromisoformat(start))
    end = request.args.get("end")
    if end:
        query = query.filter(BindingLog.bind_time <= datetime.fromisoformat(end))

    total = query.count()
    page = int(request.args.get("page", 1))
    size = int(request.args.get("size", 10))
    logs = (
        query.order_by(BindingLog.id.desc()).offset((page - 1) * size).limit(size).all()
    )
    data = [
        {
            "username": l.username,
            "student_id": l.student_id,
            "bind_time": l.bind_time.isoformat(),
            "operator": l.operator,
            "action": l.action,
        }
        for l in logs
    ]
    return jsonify({"total": total, "items": data})


@bp.route("/export/logs", methods=["GET"])
def export_logs():
    """Export logs as an Excel file."""
    logs = BindingLog.query.order_by(BindingLog.id).all()
    df = pd.DataFrame(
        [
            {
                "username": l.username,
                "student_id": l.student_id,
                "bind_time": l.bind_time,
                "operator": l.operator,
                "action": l.action,
            }
            for l in logs
        ]
    )
    from io import BytesIO

    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    filename = f"logs_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/export/accounts", methods=["GET"])
def export_accounts():
    """Export accounts as an Excel file."""
    accounts = Account.query.order_by(Account.id).all()
    df = pd.DataFrame(
        [
            {
                "username": a.username,
                "password": a.password,
                "is_bound": a.is_bound,
                "student_id": a.student_id,
                "bind_time": a.bind_time,
                "create_time": a.create_time,
            }
            for a in accounts
        ]
    )
    from io import BytesIO

    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    filename = f"accounts_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
